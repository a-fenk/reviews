from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

import time

from config import Config
from utils import lemmatize, tokenize

DUPLICATE_COLOR = 'B20000'
EMPTY_COLOR = 'A0A0A0'
FOURTH_LEVEL_COLOR = 'CFE2F3'


# def get_tags(workbook: Workbook):
#     sheet = workbook[Config.TAGS_SHEET]
#
#     columns_with_tags = []
#     tag_name_column = None
#
#     tags = {}
#
#     for raw in sheet.iter_rows():
#         tag_name = ''
#         tag_words = []
#         for cell in raw:
#             if cell.value == Config.TAG_WORD_COLUMN:
#                 columns_with_tags.append(cell.column)
#             elif cell.value == Config.TAG_NAME_COLUMN:
#                 tag_name_column = cell.column
#             elif cell.value and cell.column in columns_with_tags:
#                 tag_words.append(' '.join(lemmatize(cell.value.strip())))
#             elif cell.value and cell.column == tag_name_column:
#                 tag_name = cell.value
#         if tag_name and tag_words:
#             tags[tag_name] = tag_words
#     return tags


def get_tags(workbook: Workbook):
    sheet = workbook[Config.TAGS_SHEET]

    columns_with_tags = []
    tag_name_column = None
    tag_level_column = None

    tags = {}

    active_first_level_key = ''
    active_second_level_key = ''
    for raw in sheet.iter_rows():
        tag_words = {'words': [], 'parent': '', 'level': 0}
        tag_name = ''
        current_level = None

        for cell in raw:
            if cell.value == Config.TAG_WORD_COLUMN:
                columns_with_tags.append(cell.column)
            elif cell.value == Config.TAG_NAME_COLUMN:
                tag_name_column = cell.column
            elif cell.value == Config.TAG_LEVEL_COLUMN:
                tag_level_column = cell.column
            elif cell.value and cell.column in columns_with_tags:
                tag_words['words'].append(' '.join(lemmatize(cell.value.strip())))
            elif cell.value and cell.column == tag_name_column:
                tag_name = cell.value
            elif cell.value and cell.column == tag_level_column:
                current_level = int(cell.value)

        if tag_name and tag_words and current_level and current_level != 1:
            # if tag_name and tag_words and current_level:
            if current_level == 1:
                tag_words['level'] = current_level
                active_first_level_key = tag_name
            elif current_level == 2:
                tag_words['level'] = current_level
                tag_words['parent'] = active_first_level_key
                active_second_level_key = tag_name
            else:
                tag_words['level'] = current_level
                tag_words['parent'] = active_second_level_key

            tags[tag_name] = tag_words
    return tags


def get_masters(workbook: Workbook):
    masters = []
    sheet = workbook[Config.MASTERS_SHEET]
    for row in sheet.iter_rows():
        if row[0].value is not None and type(row[0].value) is str:
            masters.append(row[0].value.strip('/n').strip())

    return masters


def get_master_related_rows(sheet, master: str, search_range: dict, columns: list):
    results = []

    for row in sheet.iter_rows():
        search_results = []
        row_dict = {}
        for cell in row:
            column = sheet[f'{get_column_letter(cell.column)}1'].value
            if column in columns and cell.value not in columns:
                if type(cell.value) in [float, int]:
                    value = int(cell.value)
                elif type(cell.value) is str:
                    if cell.value.isdigit():
                        value = int(cell.value)
                    else:
                        value = cell.value
                elif type(cell.value) is bool:
                    value = cell.value
                elif cell.value is None:
                    value = None
                else:
                    raise TypeError(f'неподдерживаемый тип данных - "{type(cell.value)}"')
                row_dict[column] = value
            if cell.column in range(column_index_from_string(search_range['from']),
                                    column_index_from_string(search_range['to']) + 1):
                if type(cell.value) is str:
                    cell.value = cell.value.strip()
                search_results.append(cell.value)

        if row_dict and master in search_results:
            results.append(row_dict)
    return results


def merge_reviews_and_sc(reviews, all_sc, master):
    result = []
    for sc in all_sc:
        sc['added'] = False
    for review in reviews:
        for sc in all_sc:
            if review['ID container'] == sc['id container']:
                review['Address'] = sc['Address']
                review['H1-1'] = sc['H1-1']
                sc['added'] = True
                break
            else:
                review['Address'] = ''
                review['H1-1'] = ''
        result.append(review)

    for sc in all_sc:
        if not sc['added']:
            merged = {column: None for column in Config.REVIEWS_COLUMNS
                      if column not in ['ID container', 'Masters_URL',
                                        'Кол-во отзывов', 'Кол-во отзывов Corrected - TRUE']}

            merged['Masters_URL'] = master
            merged['ID container'] = sc['id container']
            merged['Address'] = sc['Address']
            merged['H1-1'] = sc['H1-1']
            merged['Кол-во отзывов'] = 0
            merged['Кол-во отзывов Corrected - TRUE'] = 0

            result.append(merged)

            # result.append({column: None for column in Config.REVIEWS_COLUMNS
            #                if column not in ['ID container', 'Masters_URL',
            #                                  'Кол-во отзывов', 'Кол-во отзывов Corrected - TRUE']} | {
            #                   'Masters_URL': master,
            #                   'ID container': sc['id container'],
            #                   'Address': sc['Address'],
            #                   'H1-1': sc['H1-1'],
            #                   'Кол-во отзывов': 0,
            #                   'Кол-во отзывов Corrected - TRUE': 0,
            #               })

    return result


def write_data_to_excel(sheet_name, workbook: Workbook, data: list, max_related_tags_by_level: dict):
    max_related_tags_by_level[1] = 1
    max_related_tags_by_level[2] = 1

    sheet = workbook.create_sheet(sheet_name)

    extra_index = 0
    for index, column in enumerate(Config.RESULT_COLUMNS):
        if column == 'name_tag':
            empty_index = 0
            for level_index, (level, value) in enumerate(max_related_tags_by_level.items()):
                for tag_index in range(0, value):
                    if level == 2:
                        sheet[f'{get_column_letter(index + tag_index + level_index + 1)}1'].value = 'совпад'
                        sheet[f'{get_column_letter(index + tag_index + level_index + 2)}1'].value = 'Обработанные'
                        sheet[f'{get_column_letter(index + tag_index + level_index + 3)}1'].value = 'НОВЫЙ ID section'
                        empty_index = 3
                    sheet[
                        f'{get_column_letter(index + tag_index + level_index + 1 + empty_index)}1'].value = f'{str(level)} уровень'
                    extra_index = tag_index
        else:
            sheet[f'{get_column_letter(index + extra_index + 1)}1'].value = column

    start_row = 2

    data = sorted(data, key=lambda x: x['Отзыв'] is None)
    for row_index, row in enumerate(data):
        extra_index = {1: 0, 2: 0, 3: 0}
        for column_index, column in enumerate(Config.RESULT_COLUMNS):
            if column == 'name_tag':
                tag_index = {1: 0, 2: 0, 3: 0}
                for tag_details in row['name_tags']:
                    if tag_details['level'] == 1:
                        additional_index = 0
                    elif tag_details['level'] == 2:
                        additional_index = max_related_tags_by_level[1] + 3
                    else:
                        additional_index = max_related_tags_by_level[1] + 3 + max_related_tags_by_level[2]

                    if 'color' in tag_details:
                        sheet[
                            f'{get_column_letter(column_index + 1 + tag_index[tag_details["level"]] + additional_index)}'
                            f'{str(start_row + row_index)}'].fill = PatternFill(patternType='solid',
                                                                                fgColor=tag_details['color'])
                    if 'tag_name' in tag_details:
                        sheet[
                            f'{get_column_letter(column_index + 1 + tag_index[tag_details["level"]] + additional_index)}' \
                            f'{str(start_row + row_index)}'].value = tag_details['tag_name']
                    extra_index[tag_details['level']] = tag_index
                    tag_index[tag_details['level']] += 1
            else:
                sheet[f'{get_column_letter(column_index + 1 + sum(extra_index.values()))}' \
                      f'{str(start_row + row_index)}'].value = row[column] if column in row else ''


def get_related_tags(review: str, tags: dict):
    related_tags = []
    for tag_name, tag_details in sorted(tags.items(), key=lambda x: x[1]['level']):
        related_tag = ''

        for tag_word in tag_details['words']:
            if review and tag_word in lemmatize(' '.join(tokenize(review))):
                # if tag_details['level'] == 1:
                #     related_tag = tag_name
                #     break

                if tag_details['level'] == 2:
                    related_tag = tag_name
                    break
                elif any([tag_details['parent'] == tag['tag_name'] for tag in related_tags if tag['level'] == 2]):
                    if tag_details['level'] != 1:
                        related_tag = tag_name
                    break
                else:
                    continue
        if related_tag:
            related_tags.append({'tag_name': related_tag, 'level': tag_details['level']})

    return related_tags


def check_is_corrected(data):
    for row in data:
        for tag in row['name_tags']:
            for original_row in data:
                if 'tag_name' in tag and tag['tag_name'] == original_row['Name section'] \
                        and row['Masters_URL'] == original_row['Masters_URL']:
                    tag['tag_name'] = f'{tag["tag_name"]} [{original_row["Кол-во отзывов Corrected - TRUE"]}]'
                    break
            if 'tag_name' in tag and '[' not in tag['tag_name'] and ']' not in tag['tag_name']:
                tag['tag_name'] = f'{tag["tag_name"]} [0]'

        row['name_tags'] = sorted(row['name_tags'], key=lambda x: int(x['tag_name'].split('[')[-1].split(']')[0] if
                                                                      'tag_name' in x else 0))
    return data


def sort_reviews():
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - STARTING')

    print(f'{time.strftime("%H:%M:%S", time.localtime())} - loading source workbook ...')
    source_wb = load_workbook(Config.SOURCE_FILE_NAME)
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - success')

    print(f'{time.strftime("%H:%M:%S", time.localtime())} - trying to get masters ...')
    masters = get_masters(source_wb)
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - success, found {len(masters)} masters')

    print(f'{time.strftime("%H:%M:%S", time.localtime())} - trying to get tags ...')
    tags = get_tags(source_wb)
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - success, found {len(tags)} tags')

    max_related_tags_by_level = {1: 0, 2: 0, 3: 0}

    result_data = []

    for master_index, master in enumerate(masters[:Config.LIMIT_MASTERS]):
        print(
            f'{time.strftime("%H:%M:%S", time.localtime())} - master = {master}, progress = {master_index + 1}/{len(masters)}')
        reviews = get_master_related_rows(
            sheet=source_wb[Config.REVIEWS_SHEET],
            master=master,
            search_range=Config.REVIEWS_SEARCH_RANGE,
            columns=Config.REVIEWS_COLUMNS,
        )
        sc = get_master_related_rows(
            sheet=source_wb[Config.SC_SHEET],
            master=master,
            search_range=Config.SC_SEARCH_RANGE,
            columns=Config.SC_COLUMNS,
        )

        reviews_and_sc = merge_reviews_and_sc(reviews, sc, master)

        for index, row in enumerate(reviews_and_sc):
            related_tags = get_related_tags(row['Отзыв'], tags)

            temp_max = {1: 0, 2: 0, 3: 0}
            for related_tag in related_tags:
                if related_tag['level'] > 3:
                    related_tag['level'] = 3
                    related_tag['color'] = FOURTH_LEVEL_COLOR

                temp_max[related_tag['level']] += 1
            for key, value in temp_max.items():
                if key == 2 and value > 1:
                    related_tags = [{'color': DUPLICATE_COLOR, 'level': 2}]
                    value = 1
                elif key == 2 and value == 0:
                    related_tags = [{'color': EMPTY_COLOR, 'level': 2}]

                if value > max_related_tags_by_level[key]:
                    max_related_tags_by_level[key] = value

            row['name_tags'] = related_tags

        result_data += check_is_corrected(reviews_and_sc)

    print(f'{time.strftime("%H:%M:%S", time.localtime())} - trying to write data to  {Config.RESULT_FILE_NAME}...')
    write_data_to_excel(
        sheet_name='result',
        workbook=source_wb,
        data=result_data,
        max_related_tags_by_level=max_related_tags_by_level
    )

    source_wb.save(Config.RESULT_FILE_NAME)
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - success')
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - FINISHED')
